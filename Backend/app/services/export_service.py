import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_excel_report(data: List[Dict[str, Any]], filters_summary: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "DealFlow360 Performance"

    # Header styling
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "DealFlow360 — Sales Performance Report"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1E3A8A")
    
    ws["A2"] = f"Filters: {filters_summary or 'All Time / All Reps'}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="64748B")

    headers = ["Quote #", "Customer", "Rep", "Total Amount ($)", "Margin %", "Risk Band", "Status"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    row_idx = 5
    for item in data:
        ws.cell(row=row_idx, column=1, value=item.get("quote_number", ""))
        ws.cell(row=row_idx, column=2, value=item.get("customer", ""))
        ws.cell(row=row_idx, column=3, value=item.get("rep", ""))
        ws.cell(row=row_idx, column=4, value=float(item.get("total_amount", 0.0)))
        ws.cell(row=row_idx, column=5, value=f"{float(item.get('margin_pct', 0.0)):.1f}%")
        ws.cell(row=row_idx, column=6, value=item.get("risk", "NONE"))
        ws.cell(row=row_idx, column=7, value=item.get("status", "").upper())

        for c in range(1, 8):
            ws.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def generate_pdf_report(data: List[Dict[str, Any]], filters_summary: str = "") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1E3A8A'),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=15
    )

    story = [
        Paragraph("DealFlow360 — Sales Operations & Performance Report", title_style),
        Paragraph(f"Active Filters: {filters_summary or 'All Time / All Reps'}", subtitle_style)
    ]

    table_data = [["Quote #", "Customer", "Sales Rep", "Amount ($)", "Margin %", "Risk", "Status"]]
    for item in data:
        table_data.append([
            str(item.get("quote_number", "")),
            str(item.get("customer", "")),
            str(item.get("rep", "")),
            f"${float(item.get('total_amount', 0.0)):,.2f}",
            f"{float(item.get('margin_pct', 0.0)):.1f}%",
            str(item.get("risk", "NONE")),
            str(item.get("status", "")).upper()
        ])

    table = Table(table_data, colWidths=[90, 140, 110, 90, 70, 70, 90])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F5F9')])
    ]))

    story.append(table)
    doc.build(story)
    return buf.getvalue()
